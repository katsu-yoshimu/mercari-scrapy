import openpyxl
from datetime import datetime

def outputExcel(key_word, data, data_detail):
    TEMPLETE_FILE_PATH = './output/template.xlsx'
    TEMPLETE_SHEET_NAME = 'メルカリ検索結果'
    OUTPUT_FILE_DIR_PATH = './output/'
    wb = ''
    ws = ''

    # Excel（テンプレート）読込
    wb = openpyxl.load_workbook(TEMPLETE_FILE_PATH)

    # シートを指定
    ws = wb[TEMPLETE_SHEET_NAME]

    # 検索キーワード
    key_word = '時計 tag heuer'
    ws.cell(row=2, column=4).value = key_word

    # 作成日時
    now = datetime.now()
    ws.cell(row=3, column=4).value = now

    # 保存先ファイル名
    output_file_path = f'{OUTPUT_FILE_DIR_PATH}output_{now.strftime("%Y%m%d_%H%M%S")}.xlsx'

    data_count = 0
    for item in data:
        data_count += 1
        row_count = data_count+6
        # 行番号
        ws.cell(row=row_count, column=2).value = data_count
        # 一覧データ出力
        ws.cell(row=row_count, column=3).value = item['mercari_id']
        ws.cell(row=row_count, column=4).value = item['name_price']
        ws.cell(row=row_count, column=5).value = item['thumb_url']
        ws.cell(row=row_count, column=6).value = item['detail_url']
        
        for detail in data_detail:
            #  "detail_url" が一致しているものがあれば
            if  item['detail_url'] == detail['detail_url']:
                # 詳細データ出力
                ws.cell(row=row_count, column=7).value = detail['name']
                ws.cell(row=row_count, column=8).value = detail['price']
                ws.cell(row=row_count, column=9).value = detail['desc'] # ★ToDo★データの先頭が「=」のときエラーになす
                col_count = 9
                for img_url in detail['img_urls']:
                    col_count += 1
                    ws.cell(row=row_count, column=col_count).value = img_url

    # Excel出力
    wb.save(output_file_path)